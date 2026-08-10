import io
from datetime import date, timedelta
from functools import wraps

from flask import (
    Blueprint, render_template, redirect, url_for, flash, request,
    abort, send_file
)
from flask_login import login_required, current_user

from extensions import db
from models import User, Project, TimeEntry
from forms import UserForm, ProjectForm

bp = Blueprint("admin", __name__, url_prefix="/admin")


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            abort(403)
        return view(*args, **kwargs)
    return wrapped


@bp.route("/")
@login_required
@admin_required
def dashboard():
    today = date.today()
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)

    week_entries = TimeEntry.query.filter(
        TimeEntry.work_date >= start_of_week, TimeEntry.work_date <= end_of_week
    ).all()

    total_users = User.query.filter_by(active=True).count()
    total_projects = Project.query.filter_by(active=True).count()
    pending = TimeEntry.query.filter_by(status="pianificato").count()

    return render_template(
        "admin/dashboard.html",
        week_entries=week_entries,
        start_of_week=start_of_week,
        end_of_week=end_of_week,
        total_users=total_users,
        total_projects=total_projects,
        pending=pending,
    )


# ---------- Gestione utenti ----------

@bp.route("/users")
@login_required
@admin_required
def users():
    all_users = User.query.order_by(User.username).all()
    return render_template("admin/users.html", users=all_users)


@bp.route("/users/new", methods=["GET", "POST"])
@login_required
@admin_required
def new_user():
    form = UserForm()
    if form.validate_on_submit():
        if User.query.filter_by(username=form.username.data.strip()).first():
            flash("Nome utente già esistente.", "danger")
        elif not form.password.data:
            flash("La password è obbligatoria per un nuovo utente.", "danger")
        else:
            user = User(
                username=form.username.data.strip(),
                full_name=form.full_name.data.strip(),
                email=form.email.data.strip() if form.email.data else None,
                role=form.role.data,
                active=form.active.data,
            )
            user.set_password(form.password.data)
            db.session.add(user)
            db.session.commit()
            flash("Utente creato.", "success")
            return redirect(url_for("admin.users"))
    return render_template("admin/user_form.html", form=form)


@bp.route("/users/<int:user_id>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    form = UserForm(obj=user)
    if request.method == "GET":
        form.password.data = ""

    if form.validate_on_submit():
        clash = User.query.filter(
            User.username == form.username.data.strip(), User.id != user.id
        ).first()
        if clash:
            flash("Nome utente già in uso da un altro account.", "danger")
        else:
            user.username = form.username.data.strip()
            user.full_name = form.full_name.data.strip()
            user.email = form.email.data.strip() if form.email.data else None
            user.role = form.role.data
            user.active = form.active.data
            if form.password.data:
                user.set_password(form.password.data)
            db.session.commit()
            flash("Utente aggiornato.", "success")
            return redirect(url_for("admin.users"))

    return render_template("admin/user_form.html", form=form, editing=True)


# ---------- Gestione progetti ----------

@bp.route("/projects")
@login_required
@admin_required
def projects():
    all_projects = Project.query.order_by(Project.name).all()
    return render_template("admin/projects.html", projects=all_projects)


@bp.route("/projects/new", methods=["GET", "POST"])
@login_required
@admin_required
def new_project():
    form = ProjectForm()
    if form.validate_on_submit():
        if Project.query.filter_by(name=form.name.data.strip()).first():
            flash("Esiste già un progetto con questo nome.", "danger")
        else:
            project = Project(
                name=form.name.data.strip(),
                description=form.description.data,
                active=form.active.data,
            )
            db.session.add(project)
            db.session.commit()
            flash("Progetto creato.", "success")
            return redirect(url_for("admin.projects"))
    return render_template("admin/project_form.html", form=form)


@bp.route("/projects/<int:project_id>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def edit_project(project_id):
    project = Project.query.get_or_404(project_id)
    form = ProjectForm(obj=project)
    if form.validate_on_submit():
        project.name = form.name.data.strip()
        project.description = form.description.data
        project.active = form.active.data
        db.session.commit()
        flash("Progetto aggiornato.", "success")
        return redirect(url_for("admin.projects"))
    return render_template("admin/project_form.html", form=form, editing=True)


# ---------- Revisione ore / export busta paga ----------

@bp.route("/entries")
@login_required
@admin_required
def entries():
    user_id = request.args.get("user_id", type=int)
    project_id = request.args.get("project_id", type=int)
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")

    query = TimeEntry.query
    if user_id:
        query = query.filter_by(user_id=user_id)
    if project_id:
        query = query.filter_by(project_id=project_id)
    if date_from:
        query = query.filter(TimeEntry.work_date >= date.fromisoformat(date_from))
    if date_to:
        query = query.filter(TimeEntry.work_date <= date.fromisoformat(date_to))

    results = query.order_by(TimeEntry.work_date.desc()).all()

    return render_template(
        "admin/entries.html",
        entries=results,
        users=User.query.order_by(User.full_name).all(),
        projects=Project.query.order_by(Project.name).all(),
        filters=request.args,
    )


@bp.route("/entries/<int:entry_id>/confirm", methods=["POST"])
@login_required
@admin_required
def confirm_entry(entry_id):
    entry = TimeEntry.query.get_or_404(entry_id)
    entry.status = "confermato"
    db.session.commit()
    flash("Voce confermata per la busta paga.", "success")
    return redirect(request.referrer or url_for("admin.entries"))


@bp.route("/entries/confirm_bulk", methods=["POST"])
@login_required
@admin_required
def confirm_bulk():
    ids = request.form.getlist("entry_ids")
    if ids:
        TimeEntry.query.filter(TimeEntry.id.in_(ids)).update(
            {"status": "confermato"}, synchronize_session=False
        )
        db.session.commit()
        flash(f"{len(ids)} voci confermate.", "success")
    return redirect(request.referrer or url_for("admin.entries"))


@bp.route("/entries/<int:entry_id>/delete", methods=["POST"])
@login_required
@admin_required
def delete_entry(entry_id):
    entry = TimeEntry.query.get_or_404(entry_id)
    db.session.delete(entry)
    db.session.commit()
    flash("Voce eliminata.", "info")
    return redirect(request.referrer or url_for("admin.entries"))


@bp.route("/export")
@login_required
@admin_required
def export():
    """Esporta in Excel (.xlsx) le ore filtrate, pronte per la busta paga."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    user_id = request.args.get("user_id", type=int)
    project_id = request.args.get("project_id", type=int)
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")

    query = TimeEntry.query
    if user_id:
        query = query.filter_by(user_id=user_id)
    if project_id:
        query = query.filter_by(project_id=project_id)
    if date_from:
        query = query.filter(TimeEntry.work_date >= date.fromisoformat(date_from))
    if date_to:
        query = query.filter(TimeEntry.work_date <= date.fromisoformat(date_to))

    results = query.order_by(TimeEntry.work_date.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ore lavorate"
    headers = ["Dipendente", "Username", "Data", "Progetto", "Ore", "Stato", "Note"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for e in results:
        ws.append([
            e.user.full_name,
            e.user.username,
            e.work_date.strftime("%d/%m/%Y"),
            e.project.name,
            e.hours,
            e.status,
            e.notes or "",
        ])

    total_row = ["", "", "", "TOTALE", sum(e.hours for e in results), "", ""]
    ws.append(total_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 3

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"ore_lavorate_{date.today().isoformat()}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
